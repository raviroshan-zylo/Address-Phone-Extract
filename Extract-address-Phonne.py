import openpyxl
import time
from selenium import webdriver
dr = webdriver.Chrome()

wb = openpyxl.load_workbook(r"C:\Users\rroshan11188\Desktop\python.xlsx")

sh = wb["Sheet1"]
dr.get(r"https://www.google.co.in/search?hl=en-IN&rlz=1C1NHXL_enIN780IN780&ei=5rykW7ScJouL0gKh5IbgCA&q=google%20cisco%20address&oq=google+cisco+address&gs_l=psy-ab.3..33i21k1.29031.49012.0.49149.21.19.0.1.1.0.340.2798.0j8j2j3.14.0....0...1c.1.64.psy-ab..6.14.3042.6..0j35i39k1j0i131i67k1j0i67k1j0i131k1j0i10k1j0i22i30k1.433.lBM_ylKhwY0&npsic=0&rflfq=1&rlha=0&rllag=28544284,77205014,13679&tbm=lcl&rldimm=9259155283852327537&lqi=ChRnb29nbGUgY2lzY28gYWRkcmVzcyIFSAGIAQE&ved=2ahUKEwj4na-z58vdAhVQFHIKHYlgAc0QvS4wAHoECAYQIw&rldoc=1&tbs=lrf:!2m4!1e17!4m2!17m1!1e2!2m1!1e3!2m1!1e16!3sIAE,lf:1,lf_ui:4#rlfi=hd:;si:9259155283852327537,l,ChRnb29nbGUgY2lzY28gYWRkcmVzcyIFSAGIAQE;mv:!3m12!1m3!1d137543.32851068632!2d77.20793259999999!3d28.561091450000003!2m3!1f0!2f0!3f0!3m2!1i421!2i471!4f13.1;tbs:lrf:!2m1!1e3!2m1!1e16!2m4!1e17!4m2!17m1!1e2!3sIAE,lf:1,lf_ui:4")

dr.find_element_by_id("lst-ib").clear()

for i in range(2, 11):
   dr.find_element_by_id("lst-ib").send_keys( sh.cell(i,3).value)
   time.sleep(1)
   dr.find_element_by_id("mKlEF").click()
   time.sleep(3)
   dr.find_element_by_id("lst-ib").clear()
   dr.find_element_by_class_name("dbg0pd").click()
   time.sleep(3)

   jnk =  dr.find_element_by_id("akp_uid_0").text
   jnk1 = jnk[jnk.find("Phone:")+6:1000]
   jnk2 = jnk1[1:jnk1.find(chr(10))]
   sh.cell(i,4).value = jnk2
   ad1 = jnk[jnk.find("Address:")+8:1000]
   ad2 =  ad1[1:ad1.find(chr(10))]   
   sh.cell(i,5).value = ad2

   
   #print (jnk)
   #print (jnkph)
   #print ( jnk[jnk.find("Phone:"):100])
   #print (jnkph[1:jnkph.find(chr(10))])
   #sh.cell(i,4).value = dr.find_element_by_class_name("LrzXr").text
   #sh.cell(i,5).value = dr.find_element_by_class_name("LrzXr zdqRlf kno-fv").text
   #sh.cell(i,6).value = dr.find_element_by_class_name("CL9Uqc ab_button").text
   






wb.save("python.xlsx")
